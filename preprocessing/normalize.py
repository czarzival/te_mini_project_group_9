import openpyxl
import numpy as np
from sklearn.preprocessing import MinMaxScaler

def load_excel_normalize(path):
    resArray = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    
    for row in sheet.iter_rows(values_only=True):
        resArray.append(row)
    
    x = np.array(resArray)
    X = []

    # Normalize each row
    for i in range(len(x)):
        newL = []
        for l in range(len(x[i]) - 1):
            if x[i][l] == '':
                continue
            newL.append(float(x[i][l]))

        # Normalize the data
        scaler = MinMaxScaler()
        normalized_L = scaler.fit_transform(np.array(newL).reshape(-1, 1)).flatten()
        X.append(normalized_L.tolist())

    return X

def write_excel(name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'A'
    
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=j + 1, column=i + 1, value=value[i][j])
    
    workbook.save(name)

path = r'C:\Users\91702\OneDrive\Desktop\bu_seg_Sept20.xlsx'
X = load_excel_normalize(path)

print(X)

write_excel('normalized_data_Spet20_7.xlsx', X)
